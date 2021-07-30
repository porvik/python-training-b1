import matplotlib.pyplot as plt
import openpyxl


class Exporter():
    @staticmethod
    def export_to_xlsx(packet_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        for c, data in enumerate(packet_data[0], 2):
            ws.cell(1, c).value = data['name']
        for r, single_packet_data in enumerate(packet_data, 2):
            ws.cell(r, 1).value = r
            for c, data in enumerate(single_packet_data, 2):
                ws.cell(r, c).value = data['value'][:-1] if isinstance(data['value'], str) else data['value']

        wb.save('test.xlsx')

class Visualizer():
    def print_text(self, packet_data):
        filtered_list = []
        for single_packet_data in packet_data:
            filtered_list.extend([data['value'] for data in single_packet_data if data['name'] == 'vibration_factor'])
        print(filtered_list)

    def print_plot(self, packet_data):
        filtered_list = []
        for single_packet_data in packet_data:
            filtered_list.extend([data['value'] for data in single_packet_data if data['name'] == 'vibration_factor'])
        print(filtered_list)
        plt.plot(filtered_list)
        plt.show()